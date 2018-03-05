# -*- coding:UTF-8 -*-
from urllib import request
from bs4 import BeautifulSoup
import re
import openpyxl

class Book(object):
	def __init__(self, id, bookName, score, count, author, press, publicationDate):
		self.id = id
		self.bookName = bookName
		self.score = score
		self.count = count
		self.author = author
		self.press = press
		self.publicationDate = publicationDate



def write07Excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '2007测试表'

    value = [["序号", "书名", "评分", "评价人数", "作者", "出版社", "出版日期"]]
    for i in range(0, 21):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")

if __name__ == "__main__":
	bookList = []
	j = 0
	while(True):
		download_url = 'https://book.douban.com/tag/%E7%BC%96%E7%A8%8B?start=' + str(j) + '&type=T'
		head = {}
		head[
			'User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36'
		download_req = request.Request(url=download_url, headers=head)
		download_response = request.urlopen(download_req)
		# 打印出来的html代码
		download_html = download_response.read().decode('utf-8', 'ignore')
		# print(download_html)
		# 通过BeautifulSoup对解析出来的html进行格式化成标签
		soup_texts = BeautifulSoup(download_html, 'lxml')
		# 存放的是页面中所有书的集合（豆瓣一页显示20条数据）
		# ul = soup_texts.find_all(name = 'li', attrs={"class": "subject-item"})
		# 查询页面标签是ul class 为subject-list 的标签
		ul = soup_texts.find_all('li', class_="subject-item")
		if(len(ul) == 0):
			break
		# print(ul[0].select('span[class="rating_nums"]')[0].string)
		# print(re.sub("\D", "", ul[0].select('span[class="pl"]')[0].string.strip()))
		# print(ul[0].select('span[class="pl"]')[0].string.strip())
		i = 1
		for li in ul:
			count = int(re.sub("\D", "", li.select('span[class="pl"]')[0].string.strip()))
			if (count > 1000):
				name = li.select('h2 > a')[0].get('title')
				score = li.select('span[class="rating_nums"]')[0].string
				message = li.select('div[class="pub"]')[0].string.strip().split('/')
				author = message[0].strip()
				press = message[2].strip()
				publicationDate = message[3].strip()
				book = Book(i, name, score, count, author, press, publicationDate)
				bookList.append(book)
				i = i + 1
		j = j + 20
	bookList = sorted(bookList, key=lambda book: book.score)
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = '2007测试表'

	value = [["序号", "书名", "评分", "评价人数", "作者", "出版社", "出版日期"]]
	for j in range(0, len(value[0])):
		sheet.cell(row=1, column=j + 1, value=str(value[0][j]))
	i = len(bookList)+1
	j = 0
	for book in bookList:
		if (j == 39):
			break
		sheet.cell(row=i, column=1, value=i-1)
		sheet.cell(row=i, column=2, value=book.bookName)
		sheet.cell(row=i, column=3, value=book.score)
		sheet.cell(row=i, column=4, value=book.count)
		sheet.cell(row=i, column=5, value=book.author)
		sheet.cell(row=i, column=6, value=book.press)
		sheet.cell(row=i, column=7, value=book.publicationDate)
		i = i-1
		j = j+1

	wb.save("abc.xlsx")
	print("写入数据成功！")
	print(len(bookList))

